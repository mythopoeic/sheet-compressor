"""Tests for the optional openpyxl adapter (SPEC §8 / PRD Seam 2).

Builds tiny .xlsx workbooks in-memory with openpyxl itself and asserts the
Grid the adapter returns (rows, origin, cellMeta, charts) — NEVER the
compression output (SPEC §8.3). The tests are skipped when openpyxl is not
installed so the pure-core conformance run stays green.
"""

import io
import os
import sys
import tempfile
import unittest

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import (
        AreaChart,
        BarChart,
        LineChart,
        PieChart,
        Reference,
        ScatterChart,
    )

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised in core-only CI
    OPENPYXL_AVAILABLE = False


# Make `sheet_compressor` importable without installation.
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PKG_ROOT = os.path.dirname(_THIS_DIR)
if _PKG_ROOT not in sys.path:
    sys.path.insert(0, _PKG_ROOT)


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
class XlsxAdapterTests(unittest.TestCase):
    # ---------------------------------------------------------------- helpers
    def _build(self, build):
        wb = Workbook()
        build(wb)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _read(self, data, **opts):
        from sheet_compressor.adapters.xlsx import read_sheet

        return read_sheet(data, opts or None)

    # ----------------------------------------------------------------- empty
    def test_empty_workbook_returns_a1_anchored_empty_grid(self):
        # An openpyxl Workbook with no cells written still has the default
        # "Sheet". Adapter should return rows=[] anchored at A1 with no
        # cellMeta (SPEC §8.1: cellMeta MAY be omitted when there is nothing
        # to describe).
        data = self._build(lambda wb: None)
        g = self._read(data)
        self.assertEqual(g["rows"], [])
        self.assertEqual(g["origin"], {"row": 1, "col": 1})
        self.assertNotIn("cellMeta", g)
        self.assertNotIn("charts", g)

    # ---------------------------------------------------- minimal grid + meta
    def test_a1_anchored_grid_and_data_types(self):
        def build(wb):
            ws = wb.active
            ws["A1"] = "Name"
            ws["B1"] = "Qty"
            ws["A2"] = "Apple"
            ws["B2"] = 3

        g = self._read(self._build(build))
        self.assertEqual(g["origin"], {"row": 1, "col": 1})
        self.assertEqual(g["rows"], [["Name", "Qty"], ["Apple", "3"]])
        self.assertEqual(
            [[m["dataType"] for m in row] for row in g["cellMeta"]],
            [["text", "text"], ["text", "number"]],
        )

    def test_true_origin_when_used_range_is_offset(self):
        def build(wb):
            ws = wb.active
            ws["C5"] = "Name"
            ws["D5"] = "Qty"
            ws["C6"] = "Apple"
            ws["D6"] = 3

        g = self._read(self._build(build))
        self.assertEqual(g["origin"], {"row": 5, "col": 3})
        self.assertEqual(g["rows"], [["Name", "Qty"], ["Apple", "3"]])

    def test_internal_gaps_are_blank_strings(self):
        # Used range A1:C2. Gaps inside surface as "" so downstream stages
        # see a rectangular grid.
        def build(wb):
            ws = wb.active
            ws["B1"] = "B1val"
            ws["A2"] = "A2val"
            ws["C2"] = 7

        g = self._read(self._build(build))
        self.assertEqual(
            g["rows"],
            [["", "B1val", ""], ["A2val", "", "7"]],
        )
        self.assertEqual(g["origin"], {"row": 1, "col": 1})
        types = [[m["dataType"] for m in row] for row in g["cellMeta"]]
        self.assertEqual(
            types,
            [["empty", "text", "empty"], ["text", "empty", "number"]],
        )

    # ---------------------------------------------------- cell type vocabulary
    def test_data_type_mapping_covers_vocabulary(self):
        def build(wb):
            ws = wb.active
            ws["A1"] = "hello"  # text
            ws["B1"] = 42  # number
            ws["C1"] = True  # bool
            ws["D1"] = "=1+1"  # formula

        g = self._read(self._build(build))
        self.assertEqual(
            [m["dataType"] for m in g["cellMeta"][0]],
            ["text", "number", "bool", "formula"],
        )

    def test_date_cells_become_date_type(self):
        from datetime import datetime

        def build(wb):
            ws = wb.active
            ws["A1"] = datetime(2024, 1, 1)

        g = self._read(self._build(build))
        self.assertEqual(g["cellMeta"][0][0]["dataType"], "date")
        # ISO-style text — exact format doesn't matter for SPEC, but adapter
        # must emit *some* string (cells are text everywhere in §1).
        self.assertTrue(isinstance(g["rows"][0][0], str))
        self.assertNotEqual(g["rows"][0][0], "")

    # ----------------------------------------------------- sheet selection
    def test_defaults_to_first_sheet(self):
        def build(wb):
            ws = wb.active
            ws["A1"] = "A"
            ws2 = wb.create_sheet("Second")
            ws2["A1"] = "B"

        g = self._read(self._build(build))
        self.assertEqual(g["rows"][0][0], "A")

    def test_select_by_name(self):
        def build(wb):
            ws = wb.active
            ws["A1"] = "A"
            ws2 = wb.create_sheet("Second")
            ws2["A1"] = "B"

        g = self._read(self._build(build), sheet="Second")
        self.assertEqual(g["rows"][0][0], "B")

    def test_select_by_index(self):
        def build(wb):
            ws = wb.active
            ws["A1"] = "A"
            ws2 = wb.create_sheet("Second")
            ws2["A1"] = "B"

        g = self._read(self._build(build), sheet=1)
        self.assertEqual(g["rows"][0][0], "B")

    def test_unknown_sheet_name_raises(self):
        data = self._build(lambda wb: wb.active.cell(row=1, column=1, value="A"))
        with self.assertRaises(ValueError):
            self._read(data, sheet="Missing")

    def test_index_out_of_range_raises(self):
        data = self._build(lambda wb: wb.active.cell(row=1, column=1, value="A"))
        with self.assertRaises(ValueError):
            self._read(data, sheet=5)

    # -------------------------------------------------------- file path entry
    def test_reads_from_file_path(self):
        def build(wb):
            wb.active["A1"] = "x"

        data = self._build(build)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(data)
            path = f.name
        try:
            from sheet_compressor.adapters.xlsx import read_sheet

            g = read_sheet(path)
            self.assertEqual(g["rows"][0][0], "x")
        finally:
            os.unlink(path)

    # --------------------------------------------------------- chart support
    def test_extracts_bar_chart_title_axes_series(self):
        def build(wb):
            ws = wb.active
            ws["A1"] = "Quarter"
            ws["B1"] = "Sales"
            ws["A2"] = "Q1"
            ws["B2"] = 100
            ws["A3"] = "Q2"
            ws["B3"] = 150
            chart = BarChart()
            chart.title = "Sales"
            chart.x_axis.title = "Quarter"
            chart.y_axis.title = "Amount"
            data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=3)
            cats = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "B5")

        g = self._read(self._build(build))
        charts = g["charts"]
        self.assertEqual(len(charts), 1)
        c = charts[0]
        self.assertEqual(c["type"], "bar")
        self.assertNotEqual(c["anchorRange"], "")
        self.assertEqual(c["title"], "Sales")
        self.assertEqual(c["axes"]["x"], "Quarter")
        self.assertEqual(c["axes"]["y"], "Amount")
        # The series name comes from B1; data range is B2:B3.
        self.assertEqual(c["series"], ["B1"])
        self.assertEqual(c["dataRanges"], ["B2:B3"])

    def test_chart_type_mapping(self):
        ctor_to_type = [
            (BarChart, "bar"),
            (LineChart, "line"),
            (PieChart, "pie"),
            (ScatterChart, "scatter"),
            (AreaChart, "area"),
        ]

        for ctor, want in ctor_to_type:
            with self.subTest(want=want):

                def build(wb, ctor=ctor):
                    ws = wb.active
                    ws["A1"] = "x"
                    ws["B1"] = 1
                    ws["B2"] = 2
                    chart = ctor()
                    if ctor is ScatterChart:
                        xvals = Reference(
                            ws, min_col=1, min_row=1, max_col=1, max_row=1
                        )
                        yvals = Reference(
                            ws, min_col=2, min_row=1, max_col=2, max_row=2
                        )
                        from openpyxl.chart import Series

                        chart.series.append(Series(yvals, xvals, title="s1"))
                    else:
                        data = Reference(
                            ws, min_col=2, min_row=1, max_col=2, max_row=2
                        )
                        chart.add_data(data)
                    ws.add_chart(chart, "D2")

                g = self._read(self._build(build))
                self.assertEqual(len(g["charts"]), 1)
                self.assertEqual(g["charts"][0]["type"], want)

    def test_no_charts_omits_charts_key(self):
        def build(wb):
            wb.active["A1"] = "x"

        g = self._read(self._build(build))
        self.assertNotIn("charts", g)

    # --------------------------------------------------------- error paths
    def test_rejects_garbage_bytes(self):
        with self.assertRaises(Exception):
            self._read(b"not a workbook")


class XlsxAdapterUnavailableTests(unittest.TestCase):
    """The adapter must surface a clear, actionable error when openpyxl is
    not installed (SPEC §8.2). Simulated by sabotaging the import.
    """

    def test_import_error_names_dependency(self):
        # Force `import openpyxl` to fail inside the adapter module by
        # blacklisting it in sys.modules.
        import importlib

        from sheet_compressor.adapters import xlsx as xlsx_mod

        saved = sys.modules.pop("openpyxl", None)
        sys.modules["openpyxl"] = None  # importing this raises ImportError
        try:
            importlib.reload(xlsx_mod)
            with self.assertRaises(ImportError) as ctx:
                xlsx_mod.read_sheet(b"\x00\x00")
            self.assertIn("openpyxl", str(ctx.exception))
        finally:
            del sys.modules["openpyxl"]
            if saved is not None:
                sys.modules["openpyxl"] = saved
            importlib.reload(xlsx_mod)


if __name__ == "__main__":
    unittest.main()
