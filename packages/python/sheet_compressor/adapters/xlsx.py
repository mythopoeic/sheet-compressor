"""Optional .xlsx adapter (SPEC §8 / PRD Seam 2) — openpyxl-backed.

Reads a single sheet — grid, origin, per-cell data type, and embedded chart
descriptors — into the core's input contract (``Grid``, see SPEC §1). The pure
compression core stays dependency-free; this module is the only place openpyxl
is referenced.

openpyxl is declared under ``[project.optional-dependencies].xlsx`` in
``pyproject.toml``. ``read_sheet`` imports it lazily and raises a clear
``ImportError`` naming the missing dependency when it's absent (SPEC §8.2).

Usage::

    from sheet_compressor import compress
    from sheet_compressor.adapters.xlsx import read_sheet

    grid = read_sheet("workbook.xlsx")
    result = compress(grid)
"""

from __future__ import annotations

import io
import os
from datetime import date, datetime, time
from typing import Any, BinaryIO, List, Mapping, Optional, Union

from ..address import a1


ReadSheetInput = Union[str, bytes, bytearray, BinaryIO, os.PathLike]


_MISSING_DEP_MSG = (
    "read_sheet() requires the optional 'openpyxl' dependency. "
    "Install it with `pip install sheet-compressor[xlsx]` or "
    "`pip install openpyxl`, or build the Grid yourself and pass it to "
    "compress()."
)


def _load_openpyxl():
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as cause:
        raise ImportError(_MISSING_DEP_MSG) from cause
    return openpyxl


# ---------------------------------------------------------------------------
# Cell value / data-type extraction
# ---------------------------------------------------------------------------


def _cell_text(value: Any) -> str:
    """Render an openpyxl cell value as a string for the Grid."""
    if value is None:
        return ""
    if isinstance(value, bool):
        # bool is a subclass of int — branch first so True/False don't get
        # stringified as "1"/"0".
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float):
        # Keep integer-valued floats compact: 3.0 → "3", matching the TS /
        # Go adapters (and SheetJS's default `w`).
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _data_type(cell) -> str:
    """Map an openpyxl ``cell.data_type`` to the SPEC §1 vocabulary.

    openpyxl uses single-letter codes: ``s`` (text), ``n`` (number),
    ``b`` (bool), ``d`` (date), ``f`` (formula), ``e`` (error),
    ``inlineStr`` / ``str`` (text). A ``cell.data_type == "n"`` with
    ``cell.value is None`` is a gap — surface as ``empty``.
    """
    t = cell.data_type
    if t == "f":
        return "formula"
    if cell.value is None:
        return "empty"
    if t in ("s", "str", "inlineStr"):
        return "text"
    if t == "n":
        return "number"
    if t == "b":
        return "bool"
    if t == "d":
        return "date"
    if t == "e":
        return "error"
    return "text"


# ---------------------------------------------------------------------------
# Chart extraction
# ---------------------------------------------------------------------------


_CHART_CLASS_TO_TYPE = {
    "BarChart": "bar",
    "BarChart3D": "bar",
    "LineChart": "line",
    "LineChart3D": "line",
    "PieChart": "pie",
    "PieChart3D": "pie",
    "DoughnutChart": "pie",
    "ScatterChart": "scatter",
    "BubbleChart": "scatter",
    "AreaChart": "area",
    "AreaChart3D": "area",
}


def _normalize_range(ref: Optional[str]) -> Optional[str]:
    """Strip sheet qualifier and absolute markers: ``'Sheet1'!$B$2:$B$4`` → ``B2:B4``."""
    if not ref:
        return None
    s = ref
    bang = s.rfind("!")
    if bang >= 0:
        s = s[bang + 1 :]
    s = s.replace("$", "")
    s = s.strip("'")
    return s


def _rich_text(title_or_axis) -> Optional[str]:
    """Concatenate every ``<a:t>`` run inside a Title block.

    openpyxl exposes axis titles and chart titles as ``Title`` objects whose
    text lives at ``title.tx.rich.p[i].r[j].t``. Either ``tx`` or ``rich`` can
    be absent; treat any missing layer as "no title".
    """
    if title_or_axis is None:
        return None
    tx = getattr(title_or_axis, "tx", None)
    if tx is None:
        return None
    rich = getattr(tx, "rich", None)
    if rich is None:
        # Could be a strRef pointing at a cell — emit the resolved ref.
        ref = getattr(tx, "strRef", None)
        if ref is not None and getattr(ref, "f", None):
            return _normalize_range(ref.f)
        return None
    parts: List[str] = []
    for para in rich.p or []:
        for run in para.r or []:
            t = getattr(run, "t", None)
            if t:
                parts.append(t)
    if not parts:
        return None
    return "".join(parts)


def _anchor_range(anchor) -> Optional[str]:
    """A1 range for the chart's anchor.

    ``TwoCellAnchor`` carries both ``_from`` and ``to`` corners. ``OneCellAnchor``
    only has ``_from`` + an ``ext`` extent — collapse to a single-cell range,
    which still satisfies SPEC §1 ("A1 range") and tells the LLM where the
    chart sits. ``AbsoluteAnchor`` (pixel-positioned) has no cell anchor at
    all — return ``None`` to skip the descriptor's ``anchorRange``.
    """
    if anchor is None:
        return None
    frm = getattr(anchor, "_from", None)
    if frm is None:
        return None
    fr = int(frm.row) + 1
    fc = int(frm.col) + 1
    to = getattr(anchor, "to", None)
    if to is not None:
        tr = int(to.row) + 1
        tc = int(to.col) + 1
        return f"{a1(fr, fc)}:{a1(tr, tc)}"
    return f"{a1(fr, fc)}:{a1(fr, fc)}"


def _series_name(series) -> Optional[str]:
    tx = getattr(series, "tx", None)
    if tx is None:
        return None
    # SeriesLabel: literal v wins over strRef.
    v = getattr(tx, "v", None)
    if v is not None:
        return str(v)
    ref = getattr(tx, "strRef", None)
    if ref is not None and getattr(ref, "f", None):
        return _normalize_range(ref.f)
    return None


def _series_data_range(series) -> Optional[str]:
    """Pull the values reference off a Series.

    Cartesian charts put the y-values on ``series.val``, scatter / bubble
    charts may keep them on ``series.yVal``; category-only charts (e.g. pie)
    fall back to ``series.cat``. We try the same order as the TS / Go ports.
    """
    for attr in ("val", "yVal", "cat"):
        block = getattr(series, attr, None)
        if block is None:
            continue
        for ref_attr in ("numRef", "strRef"):
            ref = getattr(block, ref_attr, None)
            if ref is not None and getattr(ref, "f", None):
                return _normalize_range(ref.f)
    return None


def _chart_type(chart) -> str:
    name = type(chart).__name__
    return _CHART_CLASS_TO_TYPE.get(name, "other")


def _extract_charts(ws) -> List[dict]:
    # openpyxl doesn't surface the drawing's cNvPr name on its chart objects,
    # so descriptor.name is left empty here. SPEC §1 treats `name` as
    # informational, and the chart's `title` carries the human-meaningful label.
    out: List[dict] = []
    for chart in ws._charts or []:
        anchor_range = _anchor_range(getattr(chart, "anchor", None))
        if anchor_range is None:
            continue
        descriptor: dict = {
            "name": "",
            "type": _chart_type(chart),
            "anchorRange": anchor_range,
        }
        title = _rich_text(getattr(chart, "title", None))
        if title is not None:
            descriptor["title"] = title

        axes: dict = {}
        x_title = _rich_text(getattr(getattr(chart, "x_axis", None), "title", None))
        if x_title is not None:
            axes["x"] = x_title
        y_title = _rich_text(getattr(getattr(chart, "y_axis", None), "title", None))
        if y_title is not None:
            axes["y"] = y_title
        if axes:
            descriptor["axes"] = axes

        series: List[str] = []
        data: List[str] = []
        for ser in getattr(chart, "series", None) or []:
            name = _series_name(ser)
            if name is not None:
                series.append(name)
            rng = _series_data_range(ser)
            if rng is not None:
                data.append(rng)
        if series:
            descriptor["series"] = series
        if data:
            descriptor["dataRanges"] = data
        out.append(descriptor)
    return out


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def _open_workbook(openpyxl_mod, input_: ReadSheetInput):
    """Dispatch to ``load_workbook`` with the right input shape.

    Strings / PathLike are paths on disk. Raw bytes are wrapped in a
    ``BytesIO`` because openpyxl expects a file-like object. Anything already
    file-like is passed through.
    """
    if isinstance(input_, (bytes, bytearray)):
        return openpyxl_mod.load_workbook(io.BytesIO(bytes(input_)), data_only=False)
    if isinstance(input_, (str, os.PathLike)):
        return openpyxl_mod.load_workbook(os.fspath(input_), data_only=False)
    return openpyxl_mod.load_workbook(input_, data_only=False)


def read_sheet(
    input_: ReadSheetInput,
    options: Optional[Mapping[str, Any]] = None,
) -> dict:
    """Read a single sheet from an .xlsx file and return a ``Grid`` (SPEC §1).

    ``options["sheet"]`` selects the sheet by name (``str``) or 0-indexed
    position (``int``); omitting it picks the first sheet.

    The returned dict has ``rows``, ``origin``, and — when the source has any
    cells — ``cellMeta`` parallel to ``rows``. When the worksheet has embedded
    charts, ``charts`` is set to the list of ``ChartDescriptor`` dicts in
    document order.
    """
    openpyxl = _load_openpyxl()
    opts = options or {}

    wb = _open_workbook(openpyxl, input_)
    sheet_names = wb.sheetnames
    if not sheet_names:
        raise ValueError("read_sheet(): workbook contains no sheets")

    selector = opts.get("sheet")
    if isinstance(selector, str):
        if selector not in sheet_names:
            raise ValueError(
                f"read_sheet(): sheet {selector!r} not found in workbook "
                f"(available: {', '.join(sheet_names)})"
            )
        sheet_name = selector
    else:
        idx = 0 if selector is None else int(selector)
        if idx < 0 or idx >= len(sheet_names):
            raise ValueError(
                f"read_sheet(): sheet index {idx} out of range "
                f"(workbook has {len(sheet_names)} sheet(s))"
            )
        sheet_name = sheet_names[idx]

    ws = wb[sheet_name]
    grid = _build_grid(ws)
    charts = _extract_charts(ws)
    if charts:
        grid["charts"] = charts
    return grid


def _build_grid(ws) -> dict:
    """Materialise ``ws`` into the SPEC §1 Grid shape.

    The used range is openpyxl's ``ws.dimensions``. For a sheet with no
    written cells, openpyxl reports ``"A1:A1"`` AND ``min_row==max_row==1``
    AND ``cell.value is None``; we detect that and return an empty grid
    anchored at A1 with ``cellMeta`` omitted (SPEC §8.1 lets the adapter skip
    cellMeta when there is nothing to describe).
    """
    if ws.max_row is None or ws.max_column is None:
        return {"rows": [], "origin": {"row": 1, "col": 1}}

    min_row = ws.min_row
    min_col = ws.min_column
    max_row = ws.max_row
    max_col = ws.max_column

    # Detect a truly empty sheet — openpyxl still reports min_row=max_row=1
    # for a fresh workbook with no cells.
    if (
        min_row == 1
        and max_row == 1
        and min_col == 1
        and max_col == 1
        and ws.cell(row=1, column=1).value is None
        and ws.cell(row=1, column=1).data_type == "n"
    ):
        return {"rows": [], "origin": {"row": 1, "col": 1}}

    rows: List[List[str]] = []
    cell_meta: List[List[dict]] = []
    for r in range(min_row, max_row + 1):
        row_vals: List[str] = []
        row_meta: List[dict] = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_vals.append(_cell_text(cell.value))
            row_meta.append({"dataType": _data_type(cell)})
        rows.append(row_vals)
        cell_meta.append(row_meta)

    return {
        "rows": rows,
        "origin": {"row": min_row, "col": min_col},
        "cellMeta": cell_meta,
    }
